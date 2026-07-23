from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import (
    DailyPlanItem,
    DateKind,
    LearningTask,
    MasterTask,
    NoteableType,
    Opportunity,
    OpportunityLifecycle,
    TaskStatus,
)
from app.services.daily_plan import (
    assign_for_target_date_change,
    get_or_create_daily_plan,
    set_primary_note,
)
from app.services.opportunities import (
    normalize_pipeline_stage,
    normalize_remote_status,
    normalize_source,
)


def _normalize_remote_import(value: str) -> str | None:
    return normalize_remote_status(value)


def _normalize_source_import(value: str) -> str | None:
    return normalize_source(value)


def _normalize_pipeline_import(value: str) -> str:
    return normalize_pipeline_stage(value)


def _norm_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(value) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"yes", "y", "true", "1", "x", "done", "completed"}


def _parse_int(value, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


class ImportResult:
    def __init__(self) -> None:
        self.created: dict[str, int] = {}
        self.warnings: list[str] = []

    def add(self, key: str, count: int = 1) -> None:
        self.created[key] = self.created.get(key, 0) + count

    def warn(self, message: str) -> None:
        self.warnings.append(message)


def import_workbook(
    db: Session,
    content: bytes,
    daily_plan_date: date,
    mode: str = "append",
) -> ImportResult:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    result = ImportResult()

    sheet_map = {name.strip().lower(): name for name in wb.sheetnames}

    if "master tasks" in sheet_map:
        _import_master_tasks(db, wb[sheet_map["master tasks"]], result, mode)
    if "learning tasks" in sheet_map:
        _import_learning_tasks(db, wb[sheet_map["learning tasks"]], result, mode)
    if "daily plan" in sheet_map:
        _import_daily_plan(db, wb[sheet_map["daily plan"]], daily_plan_date, result, mode)
    if "opportunities" in sheet_map:
        _import_opportunities(db, wb[sheet_map["opportunities"]], result, mode)

    db.commit()
    return result


def _import_daily_plan(
    db: Session,
    sheet,
    plan_date: date,
    result: ImportResult,
    mode: str,
) -> None:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = [_norm_header(h) for h in rows[0]]
    col = {h: i for i, h in enumerate(headers) if h}

    plan, _ = get_or_create_daily_plan(db, plan_date, assign=False)
    order = max((i.priority_order for i in plan.items), default=-1)

    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        title = _cell_str(row[col.get("goal/task", col.get("task", 0))]) if col else ""
        if not title:
            continue

        if mode == "append":
            dup = db.query(MasterTask).filter(MasterTask.task == title).first()
            if dup and any(i.master_task_id == dup.id for i in plan.items):
                result.warn(f"Daily Plan: skipped duplicate '{title}'")
                continue

        completed = _parse_bool(row[col["completed?"]]) if "completed?" in col else False
        status = TaskStatus.COMPLETED.value if completed else TaskStatus.CURRENT.value
        master = MasterTask(
            task=title,
            target_completion_date=plan_date,
            date_kind=DateKind.GOAL.value,
            status=status,
            completed_at=datetime.utcnow() if completed else None,
        )
        db.add(master)
        db.flush()

        if "notes" in col:
            notes = _cell_str(row[col["notes"]])
            if notes:
                set_primary_note(db, NoteableType.MASTER_TASK.value, master.id, notes)

        order += 1
        db.add(
            DailyPlanItem(
                daily_plan_id=plan.id,
                priority_order=order,
                master_task_id=master.id,
            )
        )
        result.add("daily_plan_items")


def _import_master_tasks(db: Session, sheet, result: ImportResult, mode: str) -> None:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = [_norm_header(h) for h in rows[0]]
    col = {h: i for i, h in enumerate(headers) if h}

    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        task_text = _cell_str(row[col.get("task", 0)])
        if not task_text:
            continue

        if mode == "append":
            dup = db.query(MasterTask).filter(MasterTask.task == task_text).first()
            if dup:
                result.warn(f"Master Tasks: skipped duplicate '{task_text}'")
                continue

        completed = _parse_bool(row[col["completed?"]]) if "completed?" in col else False
        target = (
            _parse_date(row[col["target completion date"]])
            if "target completion date" in col
            else None
        )
        master = MasterTask(
            task=task_text,
            priority=_parse_int(row[col["priority"]]) if "priority" in col else 0,
            target_completion_date=target,
            date_kind=DateKind.REQUIREMENT.value if target else DateKind.GOAL.value,
            status=TaskStatus.COMPLETED.value if completed else TaskStatus.CURRENT.value,
            completed_at=datetime.utcnow() if completed else None,
        )
        db.add(master)
        db.flush()

        if "notes" in col:
            notes = _cell_str(row[col["notes"]])
            if notes:
                set_primary_note(db, NoteableType.MASTER_TASK.value, master.id, notes)

        assign_for_target_date_change(db, target, master=master)
        result.add("master_tasks")


def _import_learning_tasks(db: Session, sheet, result: ImportResult, mode: str) -> None:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = [_norm_header(h) for h in rows[0]]
    col = {h: i for i, h in enumerate(headers) if h}

    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        task_text = _cell_str(row[col.get("task", 0)])
        if not task_text:
            continue

        if mode == "append":
            dup = db.query(LearningTask).filter(LearningTask.task == task_text).first()
            if dup:
                result.warn(f"Learning Tasks: skipped duplicate '{task_text}'")
                continue

        completed = _parse_bool(row[col["completed?"]]) if "completed?" in col else False
        target = (
            _parse_date(row[col["target completion date"]])
            if "target completion date" in col
            else None
        )
        learning = LearningTask(
            task=task_text,
            resource=_cell_str(row[col["resource"]]) if "resource" in col else None,
            priority=_parse_int(row[col["priority"]]) if "priority" in col else 0,
            target_completion_date=target,
            status=TaskStatus.COMPLETED.value if completed else TaskStatus.CURRENT.value,
            completed_at=datetime.utcnow() if completed else None,
        )
        db.add(learning)
        db.flush()

        if "notes" in col:
            notes = _cell_str(row[col["notes"]])
            if notes:
                set_primary_note(db, NoteableType.LEARNING_TASK.value, learning.id, notes)

        assign_for_target_date_change(db, target, learning=learning)
        result.add("learning_tasks")


def _import_opportunities(db: Session, sheet, result: ImportResult, mode: str) -> None:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = [_norm_header(h) for h in rows[0]]
    col = {h: i for i, h in enumerate(headers) if h}

    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        company = _cell_str(row[col.get("company", 0)])
        if not company:
            continue

        if mode == "append":
            dup = db.query(Opportunity).filter(Opportunity.company == company).first()
            if dup:
                result.warn(f"Opportunities: skipped duplicate '{company}'")
                continue

        opp = Opportunity(
            company=company,
            posting_url=_cell_str(row[col["posting url"]]) if "posting url" in col else None,
            connections=_cell_str(row[col["connections"]]) if "connections" in col else None,
            referred_by=_cell_str(row[col["referred by"]]) if "referred by" in col else None,
            location_text=_cell_str(row[col["location"]]) if "location" in col else None,
            remote_status=_normalize_remote_import(
                _cell_str(row[col["remote status"]]) if "remote status" in col else ""
            ),
            source=_normalize_source_import(
                _cell_str(row[col["source"]]) if "source" in col else ""
            ),
            stack=_cell_str(row[col["stack"]])
            if "stack" in col
            else (_cell_str(row[col["stack match"]]) if "stack match" in col else None),
            mission_fit=_cell_str(row[col["mission fit"]]) if "mission fit" in col else None,
            pipeline_stage=_normalize_pipeline_import(
                _cell_str(row[col["pipeline"]]) if "pipeline" in col else (
                    _cell_str(row[col["status"]]) if "status" in col else ""
                )
            ),
            lifecycle_status=OpportunityLifecycle.ACTIVE.value,
        )
        db.add(opp)
        db.flush()

        if "notes" in col:
            notes = _cell_str(row[col["notes"]])
            if notes:
                set_primary_note(db, NoteableType.OPPORTUNITY.value, opp.id, notes)

        result.add("opportunities")
